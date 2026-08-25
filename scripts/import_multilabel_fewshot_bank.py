"""多标签 Excel → 任务二 few-shot JSONL。

用法：
  python scripts/import_multilabel_fewshot_bank.py
  python scripts/import_multilabel_fewshot_bank.py --excel path/to.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
_SCRIPTS = Path(__file__).resolve().parent
for p in (_ROOT, _SCRIPTS):
    if str(p) not in sys.path:
        sys.path.insert(0, str(p))

from comp_paths import resolve_comp_file  # noqa: E402
from engine.classification.competition_label_map import (  # noqa: E402
    CANONICAL_CN_TAGS,
    normalize_cn_tags,
)

_DEFAULT_EXCEL_NAME = "任务二多标签few-shot扩充库.xlsx"

_TAG_SPLIT_RE = re.compile(r"[,\"'「」【】\[\]、，\s]+")


def parse_risk_tags(raw: str) -> list[str]:
    parts = [p for p in _TAG_SPLIT_RE.split(str(raw or "").strip()) if p]
    return normalize_cn_tags(parts)


def main() -> None:
    parser = argparse.ArgumentParser(description="多标签 Excel → 任务二 few-shot JSONL")
    parser.add_argument(
        "--excel",
        default=None,
        help=f"默认：配套数据/{_DEFAULT_EXCEL_NAME}（或 COMP_DATA_DIR）",
    )
    parser.add_argument(
        "--out",
        default="data/fewshot/risk_tag_fewshot_bank_multilabel.jsonl",
    )
    parser.add_argument(
        "--stats-out",
        default="data/fewshot/risk_tag_fewshot_bank_multilabel.stats.json",
    )
    parser.add_argument(
        "--review-out",
        default="data/fewshot/risk_tag_fewshot_bank_multilabel.review.md",
    )
    parser.add_argument("--min-chars", type=int, default=20, help="过短丢弃")
    args = parser.parse_args()

    import openpyxl

    excel_path = resolve_comp_file(_DEFAULT_EXCEL_NAME, explicit=args.excel)
    if not excel_path.is_file():
        raise SystemExit(f"Excel 不存在：{excel_path}")

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    # 优先「多标签样例」表
    sheet_name = next(
        (n for n in wb.sheetnames if "多标签" in n or "样例" in n),
        wb.sheetnames[0],
    )
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    col = {h: i for i, h in enumerate(headers)}
    for need in ("sample_id", "risk_tag", "violation_behavior"):
        if need not in col:
            raise SystemExit(f"缺少列 {need}，实际列={headers}")

    examples: list[dict] = []
    dropped: list[dict] = []
    seen: set[str] = set()

    for i, row in enumerate(rows[1:], 1):
        if not row:
            continue
        sid = str(row[col["sample_id"]] or "").strip()
        tag_raw = str(row[col["risk_tag"]] or "").strip()
        beh = str(row[col["violation_behavior"]] or "").strip()
        if not tag_raw or not beh:
            dropped.append({"reason": "empty", "row": i})
            continue
        tags = parse_risk_tags(tag_raw)
        if not tags:
            dropped.append({"reason": "unknown_tag", "tag": tag_raw, "row": i})
            continue
        if len(beh) < args.min_chars:
            dropped.append({"reason": "too_short", "beh": beh[:40], "row": i})
            continue
        key = re.sub(r"\s+", "", beh)
        if key in seen:
            dropped.append({"reason": "duplicate", "sample_id": sid, "row": i})
            continue
        seen.add(key)
        examples.append({
            "example_id": sid or f"FS-M{i:03d}",
            "case_id": sid or f"FS-M{i:03d}",
            "violation_behavior": beh,
            "risk_tags": tags,
            "typicality": 1.0,
            "source": excel_path.name,
        })

    out = Path(args.out)
    if not out.is_absolute():
        out = _ROOT / out
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8") as f:
        for ex in examples:
            f.write(json.dumps(ex, ensure_ascii=False) + "\n")

    per_tag = Counter(t for ex in examples for t in ex["risk_tags"])
    tag_count_dist = Counter(len(ex["risk_tags"]) for ex in examples)
    stats = {
        "examples": len(examples),
        "dropped": len(dropped),
        "drop_reasons": dict(Counter(d["reason"] for d in dropped)),
        "tags_covered": sum(1 for t in CANONICAL_CN_TAGS if per_tag[t] > 0),
        "tags_missing": [t for t in CANONICAL_CN_TAGS if per_tag[t] == 0],
        "per_tag": {t: per_tag[t] for t in CANONICAL_CN_TAGS if per_tag[t]},
        "tag_count_dist": {str(k): v for k, v in sorted(tag_count_dist.items())},
        "avg_tags_per_example": round(
            sum(len(ex["risk_tags"]) for ex in examples) / max(len(examples), 1), 3
        ),
        "avg_chars": round(
            sum(len(ex["violation_behavior"]) for ex in examples) / max(len(examples), 1), 1
        ),
        "source_excel": str(excel_path),
        "sheet": sheet_name,
    }
    stats_path = Path(args.stats_out)
    if not stats_path.is_absolute():
        stats_path = _ROOT / stats_path
    stats_path.write_text(json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8")

    review = Path(args.review_out)
    if not review.is_absolute():
        review = _ROOT / review
    lines = [
        "# 任务二多标签 few-shot 库复核",
        "",
        f"共 **{len(examples)}** 条；标签覆盖 **{stats['tags_covered']}/27**；"
        f"平均每条 **{stats['avg_tags_per_example']}** 个标签",
        "",
        f"缺类：{', '.join(stats['tags_missing']) or '无'}",
        "",
    ]
    for ex in examples:
        tags = "、".join(ex["risk_tags"])
        lines.append(f"- `{ex['example_id']}` [{tags}] {ex['violation_behavior']}")
    review.write_text("\n".join(lines), encoding="utf-8")

    print(json.dumps({
        "examples": stats["examples"],
        "dropped": stats["dropped"],
        "drop_reasons": stats["drop_reasons"],
        "tags_covered": stats["tags_covered"],
        "tags_missing": stats["tags_missing"],
        "tag_count_dist": stats["tag_count_dist"],
        "avg_tags_per_example": stats["avg_tags_per_example"],
        "avg_chars": stats["avg_chars"],
    }, ensure_ascii=False, indent=2))
    print(f"Bank   → {out}")
    print(f"Stats  → {stats_path}")
    print(f"Review → {review}")


if __name__ == "__main__":
    main()

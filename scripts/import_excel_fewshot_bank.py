"""Excel → 任务二 few-shot JSONL。

用法：
  python scripts/import_excel_fewshot_bank.py
  python scripts/import_excel_fewshot_bank.py --excel path/to.xlsx
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
_SCRIPTS = Path(__file__).resolve().parent
for p in (_ROOT, _SCRIPTS):
    if str(p) not in sys.path:
        sys.path.insert(0, str(p))

from comp_paths import resolve_comp_file  # noqa: E402
from engine.classification.competition_label_map import (  # noqa: E402
    CANONICAL_CN_TAGS,
    normalize_cn_tags,
)

_DEFAULT_EXCEL_NAME = "整合few-shot库_含MD编号标注.xlsx"

# 纯定性短句（进 Prompt 易变成「见词打标」）
_ABSTRACT_RE = re.compile(
    r"^(?:存在)?(?:"
    r"欺骗投保人|销售误导|夸大收益|虚假宣传|隐瞒与保险合同有关的重要情况|"
    r"给予投保人(?:、被保险人)?保险合同约定以外的(?:保险费回扣|其他利益)|"
    r"委托未取得合法资格的机构从事保险销售活动|"
    r"未按照规定使用经批准或者备案的保险条款、保险费率|"
    r"编制虚假财务(?:会计)?资料|对从业人员管理不到位"
    r")(?:的问题|的行为|的情况)?[。．]?$"
)


def main() -> None:
    parser = argparse.ArgumentParser(description="Excel → 任务二 few-shot JSONL")
    parser.add_argument(
        "--excel",
        default=None,
        help=f"默认：配套数据/{_DEFAULT_EXCEL_NAME}（或 COMP_DATA_DIR）",
    )
    parser.add_argument("--out", default="data/fewshot/risk_tag_fewshot_bank_excel.jsonl")
    parser.add_argument("--stats-out", default="data/fewshot/risk_tag_fewshot_bank_excel.stats.json")
    parser.add_argument("--review-out", default="data/fewshot/risk_tag_fewshot_bank_excel.review.md")
    parser.add_argument("--min-chars", type=int, default=8, help="过短丢弃")
    parser.add_argument(
        "--keep-abstract",
        action="store_true",
        help="保留纯定性短句（默认过滤）",
    )
    args = parser.parse_args()

    import openpyxl

    excel_path = resolve_comp_file(_DEFAULT_EXCEL_NAME, explicit=args.excel)
    if not excel_path.is_file():
        raise SystemExit(f"找不到 Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    col = {h: i for i, h in enumerate(headers)}
    for need in ("risk_tag", "违法行为", "file_id"):
        if need not in col:
            raise SystemExit(f"缺少列 {need}，实际列={headers}")

    examples: list[dict] = []
    dropped: list[dict] = []
    seen: set[tuple[str, str]] = set()

    for i, row in enumerate(rows[1:], 1):
        if not row:
            continue
        tag_raw = str(row[col["risk_tag"]] or "").strip()
        beh = str(row[col["违法行为"]] or "").strip()
        fid = str(row[col["file_id"]] or "").strip()
        if not tag_raw or not beh:
            dropped.append({"reason": "empty", "row": i})
            continue
        tags = normalize_cn_tags([tag_raw])
        if not tags:
            dropped.append({"reason": "unknown_tag", "tag": tag_raw, "row": i})
            continue
        if len(beh) < args.min_chars:
            dropped.append({"reason": "too_short", "beh": beh, "row": i})
            continue
        if not args.keep_abstract and _ABSTRACT_RE.match(beh):
            dropped.append({"reason": "abstract", "tag": tag_raw, "beh": beh, "row": i})
            continue
        key = (tags[0], re.sub(r"\s+", "", beh))
        if key in seen:
            dropped.append({"reason": "duplicate", "tag": tags[0], "beh": beh, "row": i})
            continue
        seen.add(key)
        examples.append({
            "example_id": f"EX_{fid or i:0>6}_{tags[0]}",
            "case_id": fid,  # 外库用 file_id；与评测 case_id 不重叠即可防自泄漏
            "file_id": fid,
            "violation_behavior": beh,
            "risk_tags": tags,
            "typicality": 1.0,
            "source": "整合few-shot库_含MD编号标注.xlsx",
        })

    out = Path(args.out)
    if not out.is_absolute():
        out = _ROOT / out
    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8") as f:
        for ex in examples:
            f.write(json.dumps(ex, ensure_ascii=False) + "\n")

    per_tag = Counter(t for ex in examples for t in ex["risk_tags"])
    stats = {
        "examples": len(examples),
        "dropped": len(dropped),
        "drop_reasons": dict(Counter(d["reason"] for d in dropped)),
        "tags_covered": sum(1 for t in CANONICAL_CN_TAGS if per_tag[t] > 0),
        "tags_missing": [t for t in CANONICAL_CN_TAGS if per_tag[t] == 0],
        "per_tag": {t: per_tag[t] for t in CANONICAL_CN_TAGS if per_tag[t]},
        "avg_chars": round(
            sum(len(ex["violation_behavior"]) for ex in examples) / max(len(examples), 1), 1
        ),
        "source_excel": str(excel_path),
    }
    stats_path = Path(args.stats_out)
    if not stats_path.is_absolute():
        stats_path = _ROOT / stats_path
    stats_path.write_text(json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8")

    review = Path(args.review_out)
    if not review.is_absolute():
        review = _ROOT / review
    lines = [
        "# Excel 任务二 few-shot 库复核",
        "",
        f"共 **{len(examples)}** 条（过滤后）；缺类：{', '.join(stats['tags_missing']) or '无'}",
        "",
    ]
    by_tag: dict[str, list[dict]] = {}
    for ex in examples:
        by_tag.setdefault(ex["risk_tags"][0], []).append(ex)
    for tag in CANONICAL_CN_TAGS:
        items = by_tag.get(tag) or []
        if not items:
            continue
        lines.append(f"## {tag}（{len(items)}）")
        lines.append("")
        for ex in items:
            lines.append(f"- `{ex['file_id']}` {ex['violation_behavior']}")
        lines.append("")
    review.write_text("\n".join(lines), encoding="utf-8")

    print(json.dumps({
        "examples": stats["examples"],
        "dropped": stats["dropped"],
        "drop_reasons": stats["drop_reasons"],
        "tags_covered": stats["tags_covered"],
        "tags_missing": stats["tags_missing"],
        "avg_chars": stats["avg_chars"],
    }, ensure_ascii=False, indent=2))
    print(f"Bank   → {out}")
    print(f"Stats  → {stats_path}")
    print(f"Review → {review}")


if __name__ == "__main__":
    main()
